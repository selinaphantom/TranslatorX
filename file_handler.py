#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FileHandler — 處理 CSV / Excel / TXT 檔案翻譯
"""

import csv
import time
from pathlib import Path
from typing import Callable, List, Optional


class FileHandler:
    def translate_file(
        self,
        file_path: str,
        src: str,
        tgt: str,
        api: str,
        engine,
        skip_header: bool = True,
        columns: Optional[List[str]] = None,
        delay: float = 0.3,
        progress_cb: Optional[Callable] = None,
        log_cb: Optional[Callable]      = None,
        stop_flag: Optional[Callable]   = None,
    ) -> str:
        path = Path(file_path)
        ext  = path.suffix.lower()

        if ext == ".txt":
            return self._translate_txt(path, src, tgt, api, engine,
                                       delay, progress_cb, log_cb, stop_flag)
        elif ext == ".csv":
            return self._translate_csv(path, src, tgt, api, engine,
                                       skip_header, columns, delay,
                                       progress_cb, log_cb, stop_flag)
        elif ext in (".xlsx", ".xls"):
            return self._translate_excel(path, src, tgt, api, engine,
                                         skip_header, columns, delay,
                                         progress_cb, log_cb, stop_flag)
        else:
            raise ValueError(f"不支援的檔案格式：{ext}")

    # ── TXT ───────────────────────────────────────────────────────────────
    def _translate_txt(self, path, src, tgt, api, engine,
                        delay, progress_cb, log_cb, stop_flag):
        text = path.read_text(encoding="utf-8", errors="replace")
        lines = text.splitlines()
        total = len(lines)
        out_lines = []

        for i, line in enumerate(lines):
            if stop_flag and stop_flag():
                break
            if line.strip():
                try:
                    result = engine.translate(line, src, tgt, api)
                    out_lines.append(result)
                    if log_cb:
                        log_cb(f"[{i+1}/{total}] {line[:30]}... → {result[:30]}", "info")
                except Exception as e:
                    out_lines.append(line)
                    if log_cb:
                        log_cb(f"錯誤 第{i+1}行：{e}", "error")
            else:
                out_lines.append(line)

            if progress_cb:
                progress_cb(i + 1, total, f"第 {i+1} 行")
            if delay > 0 and line.strip():
                time.sleep(delay)

        out_path = path.parent / f"{path.stem}_translated{path.suffix}"
        out_path.write_text("\n".join(out_lines), encoding="utf-8")
        return str(out_path)

    # ── CSV ───────────────────────────────────────────────────────────────
    def _translate_csv(self, path, src, tgt, api, engine,
                        skip_header, columns, delay,
                        progress_cb, log_cb, stop_flag):
        # 偵測編碼
        encoding = self._detect_encoding(path)

        with open(path, "r", encoding=encoding, errors="replace", newline="") as f:
            reader = list(csv.reader(f))

        if not reader:
            raise ValueError("CSV 檔案為空")

        header = reader[0] if skip_header else None
        data   = reader[1:] if skip_header else reader

        # 解析欄位索引
        col_indices = self._resolve_columns(columns, header, len(reader[0]))

        total = len(data)
        out_rows = []

        if header:
            out_rows.append(header)

        for i, row in enumerate(data):
            if stop_flag and stop_flag():
                break

            new_row = list(row)
            for idx in col_indices:
                if idx < len(row) and row[idx].strip():
                    try:
                        new_row[idx] = engine.translate(row[idx], src, tgt, api)
                        if log_cb:
                            log_cb(f"[{i+1}/{total}] 欄{idx+1}: {row[idx][:25]}...", "info")
                        if delay > 0:
                            time.sleep(delay)
                    except Exception as e:
                        if log_cb:
                            log_cb(f"錯誤 第{i+1}列欄{idx+1}：{e}", "error")

            out_rows.append(new_row)
            if progress_cb:
                progress_cb(i + 1, total, f"第 {i+1} 列")

        out_path = path.parent / f"{path.stem}_translated{path.suffix}"
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(out_rows)

        return str(out_path)

    # ── Excel ─────────────────────────────────────────────────────────────
    def _translate_excel(self, path, src, tgt, api, engine,
                          skip_header, columns, delay,
                          progress_cb, log_cb, stop_flag):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("請安裝 openpyxl：pip install openpyxl")

        wb = openpyxl.load_workbook(path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows())
            if not rows:
                continue

            start_row = 1 if skip_header else 0
            header_row = rows[0] if skip_header else None
            header_vals = [str(c.value or "") for c in header_row] if header_row else []
            data_rows = rows[start_row:]

            max_col = ws.max_column
            col_indices = self._resolve_columns(columns, header_vals, max_col)

            total = len(data_rows)
            for i, row in enumerate(data_rows):
                if stop_flag and stop_flag():
                    break

                for idx in col_indices:
                    if idx < len(row):
                        cell = row[idx]
                        val  = str(cell.value or "").strip()
                        if val:
                            try:
                                cell.value = engine.translate(val, src, tgt, api)
                                if log_cb:
                                    log_cb(f"[{sheet_name}] 第{i+1}列欄{idx+1}: {val[:20]}...", "info")
                                if delay > 0:
                                    time.sleep(delay)
                            except Exception as e:
                                if log_cb:
                                    log_cb(f"錯誤：{e}", "error")

                if progress_cb:
                    progress_cb(i + 1, total, f"工作表「{sheet_name}」第 {i+1} 列")

        out_path = path.parent / f"{path.stem}_translated{path.suffix}"
        wb.save(out_path)
        return str(out_path)

    # ── 工具方法 ──────────────────────────────────────────────────────────
    def _resolve_columns(self, columns, header, max_col) -> List[int]:
        """解析欄位設定，回傳 0-based 索引列表"""
        if not columns:
            return list(range(max_col))

        indices = []
        for c in columns:
            c = c.strip()
            # 字母欄位 (A, B, C...)
            if c.isalpha():
                idx = ord(c.upper()) - ord("A")
                indices.append(idx)
            # 數字欄位 (1, 2, 3...)
            elif c.isdigit():
                indices.append(int(c) - 1)
            # 欄位名稱
            elif header and c in header:
                indices.append(header.index(c))

        return indices or list(range(max_col))

    def _detect_encoding(self, path: Path) -> str:
        """簡單偵測 CSV 編碼"""
        for enc in ("utf-8-sig", "utf-8", "big5", "gbk", "cp1252"):
            try:
                with open(path, "r", encoding=enc) as f:
                    f.read(1024)
                return enc
            except (UnicodeDecodeError, LookupError):
                continue
        return "utf-8"
